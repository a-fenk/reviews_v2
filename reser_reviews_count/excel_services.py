from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

import time

from .config import Config

NEW_ID_CONTAINER_COLOR = '228B22'
GEO_COLOR = 'FF0000'


def write_data_to_excel(sheet_name, workbook: Workbook, data: dict):
    sheet = workbook.copy_worksheet(workbook[Config.REVIEWS_SHEET])
    sheet.title = sheet_name
    columns = {}
    for row in sheet.iter_rows(max_row=1):
        for cell in row:
            if cell.value == Config.GEO_METRO_COLUMN:
                columns[Config.GEO_METRO_COLUMN] = cell.column
            elif cell.value == Config.GEO_DISTRICT_COLUMN:
                columns[Config.GEO_DISTRICT_COLUMN] = cell.column
            elif cell.value == Config.REVIEWS_COUNT_COLUMN:
                columns[Config.REVIEWS_COUNT_COLUMN] = cell.column
            elif cell.value == Config.ID_SECTION_COLUMN:
                columns[Config.ID_SECTION_COLUMN] = cell.column
            elif cell.value == Config.MASTER_COLUMN:
                columns[Config.MASTER_COLUMN] = cell.column
            elif cell.value == Config.ID_CONTAINER_COLUMN:
                columns[Config.ID_CONTAINER_COLUMN] = cell.column
            elif cell.value == Config.NEW_ID_CONTAINER_COLUMN:
                columns[Config.NEW_ID_CONTAINER_COLUMN] = cell.column

    for row, values in data.items():
        for key, value in values.items():
            orig_data = sheet[f'{get_column_letter(columns[key])}{row}'].value
            sheet[f'{get_column_letter(columns[key])}{row}'] = value
            if orig_data is None and value:
                if key == Config.NEW_ID_CONTAINER_COLUMN:
                    sheet[f'{get_column_letter(columns[key])}{row}'].fill = PatternFill(patternType='solid',
                                                                                        fgColor=NEW_ID_CONTAINER_COLOR)
                if (key == Config.GEO_METRO_COLUMN or key == Config.GEO_DISTRICT_COLUMN) \
                        and Config.NEW_ID_CONTAINER_COLUMN in values and values[Config.NEW_ID_CONTAINER_COLUMN]:
                    sheet[f'{get_column_letter(columns[key])}{row}'].fill = PatternFill(patternType='solid',
                                                                                        fgColor=GEO_COLOR)


def get_data(sheet, geo: dict):
    id_section_column, master_column, reviews_count_column, metro_column, district_column, id_container_column = \
        None, None, None, None, None, None

    data = {}

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value == Config.GEO_METRO_COLUMN:
                metro_column = cell.column
            elif cell.value == Config.GEO_DISTRICT_COLUMN:
                district_column = cell.column
            elif cell.value == Config.REVIEWS_COUNT_COLUMN:
                reviews_count_column = cell.column
            elif cell.value == Config.ID_SECTION_COLUMN:
                id_section_column = cell.column
            elif cell.value == Config.MASTER_COLUMN:
                master_column = cell.column
            elif cell.value == Config.H1_COLUMN:
                h1_column = cell.column
            elif cell.value == Config.ID_CONTAINER_COLUMN:
                id_container_column = cell.column

            elif cell.column == metro_column:
                if cell.row not in data:
                    data[cell.row] = {}
                data[cell.row][Config.GEO_METRO_COLUMN] = cell.value
                if Config.GEO_DISTRICT_COLUMN in data[cell.row]:
                    geo_data = data[cell.row][Config.GEO_DISTRICT_COLUMN]
                elif Config.GEO_METRO_COLUMN in data[cell.row]:
                    geo_data = data[cell.row][Config.GEO_METRO_COLUMN]
                if not geo_data:
                    for geolocation in geo[Config.GEO_METRO_COLUMN]:
                        if sheet[f'{get_column_letter(h1_column)}{cell.row}'].value and \
                                geolocation.lower() in sheet[f'{get_column_letter(h1_column)}{cell.row}'].value.lower():
                            data[cell.row][Config.GEO_METRO_COLUMN] = geolocation
                            break
            elif cell.column == district_column:
                if cell.row not in data:
                    data[cell.row] = {}
                data[cell.row][Config.GEO_DISTRICT_COLUMN] = cell.value
                geo_data = None
                if Config.GEO_DISTRICT_COLUMN in data[cell.row]:
                    geo_data = data[cell.row][Config.GEO_DISTRICT_COLUMN]
                elif Config.GEO_METRO_COLUMN in data[cell.row]:
                    geo_data = data[cell.row][Config.GEO_METRO_COLUMN]
                if not geo_data:
                    for geolocation in geo[Config.GEO_DISTRICT_COLUMN]:
                        if sheet[f'{get_column_letter(h1_column)}{cell.row}'].value and \
                                geolocation in sheet[f'{get_column_letter(h1_column)}{cell.row}'].value:
                            data[cell.row][Config.GEO_DISTRICT_COLUMN] = geolocation
                            break
            elif cell.column == master_column:
                if cell.row not in data:
                    data[cell.row] = {}
                data[cell.row][Config.MASTER_COLUMN] = cell.value
            elif cell.column == id_section_column:
                if cell.row not in data:
                    data[cell.row] = {}
                data[cell.row][Config.ID_SECTION_COLUMN] = cell.value
            elif cell.column == reviews_count_column:
                if cell.row not in data:
                    data[cell.row] = {}
                data[cell.row][Config.REVIEWS_COUNT_COLUMN] = int(cell.value) if cell.value else 0
            elif cell.column == id_container_column:
                if cell.row not in data:
                    data[cell.row] = {}
                data[cell.row][Config.ID_CONTAINER_COLUMN] = cell.value
    return data


def get_geo(sheet):
    res = {Config.GEO_DISTRICT_COLUMN: [], Config.GEO_METRO_COLUMN: []}
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value not in [Config.GEO_DISTRICT_COLUMN, Config.GEO_METRO_COLUMN, None]:
                res[sheet[f'{get_column_letter(cell.column)}1'].value].append(str(cell.value))
    return res


def reset_reviews_count():
    print(f'{time.strftime("%H:%M:%S", time.localtime())} - STARTING')

    print(f'{time.strftime("%H:%M:%S", time.localtime())} - loading source workbook ...')
    source_wb = load_workbook(Config.SOURCE_FILE)
    print(f'{time.strftime("%H:%M:%S", time.localtime())} - success')

    print(f'{time.strftime("%H:%M:%S", time.localtime())} - loading geo from workbook ...')
    geo = get_geo(sheet=source_wb[Config.GEO_SHEET])
    print(f'{time.strftime("%H:%M:%S", time.localtime())} - success')

    print(f'{time.strftime("%H:%M:%S", time.localtime())} - getting data from workbook ...')
    data = get_data(sheet=source_wb[Config.REVIEWS_SHEET], geo=geo)
    print(f'{time.strftime("%H:%M:%S", time.localtime())} - success')

    less = {key: value for key, value in data.items() if value[Config.REVIEWS_COUNT_COLUMN] <= Config.LTE and
            (value[Config.GEO_METRO_COLUMN] or value[Config.GEO_DISTRICT_COLUMN])}
    greater = {key: value for key, value in data.items() if value[Config.REVIEWS_COUNT_COLUMN] >= Config.GTE and
               not value[Config.GEO_METRO_COLUMN] and not value[Config.GEO_DISTRICT_COLUMN]}
    # print('--------------------------------------------------------------------')
    # print(less)

    for less_elem in less.values():
        for greater_elem in greater.values():
            if greater_elem[Config.REVIEWS_COUNT_COLUMN] > Config.TARGET:
                if less_elem[Config.MASTER_COLUMN] == greater_elem[Config.MASTER_COLUMN] and \
                        less_elem[Config.ID_SECTION_COLUMN] == greater_elem[Config.ID_SECTION_COLUMN]:
                    can_take = 1 if greater_elem[Config.REVIEWS_COUNT_COLUMN] - 1 >= Config.TARGET else 0

                    # greater_elem[Config.REVIEWS_COUNT_COLUMN] -= can_take
                    # less_elem[Config.REVIEWS_COUNT_COLUMN] += can_take

                    if can_take > 0:
                        less_elem[Config.REVIEWS_COUNT_COLUMN] += can_take
                        greater_elem[Config.NEW_ID_CONTAINER_COLUMN] = less_elem[Config.ID_CONTAINER_COLUMN]
                        if less_elem[Config.GEO_DISTRICT_COLUMN]:
                            greater_elem[Config.GEO_DISTRICT_COLUMN] = less_elem[Config.GEO_DISTRICT_COLUMN]
                        else:
                            greater_elem[Config.GEO_METRO_COLUMN] = less_elem[Config.GEO_METRO_COLUMN]

                        for second_greater_elem in greater.values():
                            if less_elem[Config.MASTER_COLUMN] == second_greater_elem[Config.MASTER_COLUMN] and \
                                    (less_elem[Config.ID_SECTION_COLUMN] == second_greater_elem[
                                        Config.ID_SECTION_COLUMN]):
                                if Config.NEW_ID_CONTAINER_COLUMN in second_greater_elem and \
                                        second_greater_elem[Config.NEW_ID_CONTAINER_COLUMN] == \
                                        less_elem[Config.ID_CONTAINER_COLUMN] and \
                                        (less_elem[Config.GEO_METRO_COLUMN] == second_greater_elem[Config.GEO_METRO_COLUMN]
                                    or less_elem[Config.GEO_DISTRICT_COLUMN] == second_greater_elem[
                                        Config.GEO_DISTRICT_COLUMN]) and (
                                        second_greater_elem[Config.GEO_METRO_COLUMN] or
                                        second_greater_elem[Config.GEO_DISTRICT_COLUMN]):
                                    second_greater_elem[Config.REVIEWS_COUNT_COLUMN] = less_elem[
                                        Config.REVIEWS_COUNT_COLUMN]
                                elif not (second_greater_elem[Config.GEO_DISTRICT_COLUMN] or
                                          second_greater_elem[Config.GEO_METRO_COLUMN]):
                                    second_greater_elem[Config.REVIEWS_COUNT_COLUMN] -= can_take

                    if less_elem[Config.REVIEWS_COUNT_COLUMN] == Config.TARGET:
                        break
    # print('--------------------------------------------------------------------')
    # print(less)
    # print('--------------------------------------------------------------------')

    print(f'{time.strftime("%H:%M:%S", time.localtime())} - trying to write data to  {Config.RESULT_FILE}...')
    write_data_to_excel(
        sheet_name='corrected',
        workbook=source_wb,
        data=less | greater,
    )
    source_wb.save(Config.RESULT_FILE)
    print(f'{time.strftime("%H:%M:%S", time.localtime())} - success')

    print(f'{time.strftime("%H:%M:%S", time.localtime())} - FINISHED')
