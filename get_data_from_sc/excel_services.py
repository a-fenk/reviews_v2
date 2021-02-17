from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, column_index_from_string

import time

from .config import Config
from .utils import lemmatize, tokenize

DUPLICATE_COLOR = 'B20000'
EMPTY_COLOR = 'A0A0A0'
FOURTH_LEVEL_COLOR = 'CFE2F3'


def get_tags(workbook: Workbook):
    sheet = workbook[Config.TAGS_SHEET]

    columns_with_tags = []
    tag_name_column = None
    tag_level_column = None

    tags = {}

    active_first_level_key = ''
    active_second_level_key = ''
    for raw in sheet.iter_rows():
        tag_words = {'words': [], 'parent': '', 'level': 0}
        tag_name = ''
        current_level = None

        for cell in raw:
            if cell.value == Config.TAG_WORD_COLUMN:
                columns_with_tags.append(cell.column)
            elif cell.value == Config.TAG_NAME_COLUMN:
                tag_name_column = cell.column
            elif cell.value == Config.TAG_LEVEL_COLUMN:
                tag_level_column = cell.column
            elif cell.value and cell.column in columns_with_tags:
                tag_words['words'].append(' '.join(lemmatize(cell.value.strip())))
            elif cell.value and cell.column == tag_name_column:
                tag_name = cell.value
            elif cell.value and cell.column == tag_level_column:
                current_level = int(cell.value)

        if tag_name and tag_words and current_level and current_level != 1:
            # if tag_name and tag_words and current_level:
            if current_level == 1:
                tag_words['level'] = current_level
                active_first_level_key = tag_name
            elif current_level == 2:
                tag_words['level'] = current_level
                tag_words['parent'] = active_first_level_key
                active_second_level_key = tag_name
            else:
                tag_words['level'] = current_level
                tag_words['parent'] = active_second_level_key

            tags[tag_name] = tag_words
    return tags


def get_masters(workbook: Workbook):
    masters = []
    sheet = workbook[Config.REVIEWS_SHEET]
    for row in sheet.iter_rows(min_row=2):
        if row[1].value is not None and type(row[1].value) is str and row[1].value.strip('/n').strip() not in masters:
            masters.append(row[1].value.strip('/n').strip())

    return masters


def get_master_related_rows(sheet, master: str, search_range: dict, columns: list):
    results = []

    for row in sheet.iter_rows():
        search_results = []
        row_dict = {}
        for cell in row:
            column = sheet[f'{get_column_letter(cell.column)}1'].value
            if column in columns and cell.value not in columns:
                if type(cell.value) in [float, int]:
                    value = int(cell.value)
                elif type(cell.value) is str:
                    if cell.value.isdigit():
                        value = int(cell.value)
                    else:
                        value = cell.value
                elif type(cell.value) is bool:
                    value = cell.value
                elif cell.value is None:
                    value = None
                else:
                    raise TypeError(f'неподдерживаемый тип данных - "{type(cell.value)}"')
                row_dict[column] = value
            if cell.column in range(column_index_from_string(search_range['from']),
                                    column_index_from_string(search_range['to']) + 1):
                if type(cell.value) is str:
                    cell.value = cell.value.strip()
                search_results.append(cell.value)

        if row_dict and master in search_results:
            results.append(row_dict)
    return results


def merge_reviews_and_sc(reviews, all_sc, master):
    result = []
    for sc in all_sc:
        sc['added'] = False
    for review in reviews:
        for sc in all_sc:
            if review['ID container'] == sc['id container']:
                review['Address'] = sc['Address']
                review['H1-1'] = sc['H1-1']
                if sc['Id section 1']:
                    review['ID section'] = sc['Id section 1']
                sc['added'] = True
                break
            else:
                review['Address'] = ''
                review['H1-1'] = ''
                if 'ID section' not in review:
                    review['ID section'] = ''
        result.append(review)

    for sc in all_sc:
        if not sc['added']:
            merged = {column: None for column in Config.RESULT_COLUMNS}

            merged['Masters_URL'] = master
            merged['ID container'] = sc['id container']
            merged['Address'] = sc['Address']
            merged['H1-1'] = sc['H1-1']
            merged['ID section'] = sc['Id section 1']
            merged['Кол-во отзывов'] = 0
            merged['Кол-во отзывов Corrected - TRUE'] = 0

            result.append(merged)
    return result


def write_data_to_excel(sheet_name, workbook: Workbook, data: list):
    sheet = workbook[sheet_name] if sheet_name in workbook.sheetnames else workbook.create_sheet(sheet_name)

    start_row = 2

    data = sorted(data, key=lambda x: x['Отзыв'] is None)
    for row_index, row in enumerate(data):
        for column_index, column in enumerate(Config.RESULT_COLUMNS):
            sheet[f'{get_column_letter(column_index + 1)}{str(start_row + row_index)}'].value = row[column] \
                if column in row else ''


def sort_reviews():
    print(f'{time.strftime("%H:%M:%S", time.localtime())} - STARTING')

    print(f'{time.strftime("%H:%M:%S", time.localtime())} - loading source workbook ...')
    source_wb = load_workbook(Config.SOURCE_FILE_NAME)
    print(f'{time.strftime("%H:%M:%S", time.localtime())} - success')

    print(f'{time.strftime("%H:%M:%S", time.localtime())} - trying to get masters ...')
    masters = get_masters(source_wb)
    masters = [master for master in masters if master == 'mitskevich-ea']
    print(f'{time.strftime("%H:%M:%S", time.localtime())} - success, found {len(masters)} masters')

    result_data = []

    for master_index, master in enumerate(masters[:Config.LIMIT_MASTERS]):
        print(
            f'{time.strftime("%H:%M:%S", time.localtime())} - master = {master}, progress = {master_index + 1}/{len(masters[:Config.LIMIT_MASTERS])}')
        reviews = get_master_related_rows(
            sheet=source_wb[Config.REVIEWS_SHEET],
            master=master,
            search_range=Config.REVIEWS_SEARCH_RANGE,
            columns=Config.REVIEWS_COLUMNS,
        )
        sc = get_master_related_rows(
            sheet=source_wb[Config.SC_SHEET],
            master=master,
            search_range=Config.SC_SEARCH_RANGE,
            columns=Config.SC_COLUMNS,
        )
        result_data += merge_reviews_and_sc(reviews, sc, master)

    print(f'{time.strftime("%H:%M:%S", time.localtime())} - trying to write data to  {Config.RESULT_FILE_NAME}...')
    write_data_to_excel(
        sheet_name=Config.REVIEWS_SHEET,
        workbook=source_wb,
        data=result_data,
    )
    source_wb.save(Config.RESULT_FILE_NAME)
    print(f'{time.strftime("%H:%M:%S", time.localtime())} - success')
    print(f'{time.strftime("%H:%M:%S", time.localtime())} - FINISHED')
